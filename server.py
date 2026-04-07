"""
FastAPI backend for the TaxoConf Workspace.

Serves the index.html frontend and exposes JSON API endpoints that
the frontend calls for:
  - Oral session organization   (GET /api/oral/info, POST /api/oral/run)
  - Poster session organization (GET /api/poster/info, POST /api/poster/run)
  - Paper assignment            (placeholder)
  - Reviewer discovery          (placeholder)

Usage:
    pip install fastapi uvicorn pyyaml scikit-learn
    python server.py                     # starts on http://127.0.0.1:8000
    python server.py --port 9000         # custom port
    python server.py --host 0.0.0.0      # bind all interfaces

The server auto-discovers conference datasets under ./data/<conference_name>/
and builds taxonomy + similarity on first request (cached for subsequent calls).
"""

import argparse
import json
import logging
import os
import sys
import traceback
from collections import defaultdict
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import uvicorn

# ── Add project root to path so we can import our modules ──
PROJECT_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(PROJECT_ROOT))

import config
from models import Paper, TaxonomyNode, Session, PosterSession, FloorPlanType
from llm_client import LLMClient
from taxonomy_builder import (TaxonomyBuilder, collect_leaves,
                              print_taxonomy)
from session_organizer import run_oral_organization, OrganizationResult
from poster_organizer import run_poster_pipeline, PosterOrganizationResult
from session_namer import name_sessions
from similarity import SimilarityEngine
from token_tracker import get_global_tracker, reset_global_tracker
from session_reviewer import review_sessions

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("server")


# ════════════════════════════════════════════════════════════════════
# Data discovery and caching
# ════════════════════════════════════════════════════════════════════

DATA_DIR = PROJECT_ROOT / "data"


def discover_conferences() -> list[str]:
    """Find available conference datasets under ./data/."""
    if not DATA_DIR.is_dir():
        return []
    conferences = []
    for entry in sorted(DATA_DIR.iterdir()):
        if entry.is_dir() and not entry.name.startswith("."):
            # Check for any JSON file in the directory
            json_files = list(entry.glob("*.json"))
            if json_files:
                conferences.append(entry.name)
    return conferences


def _resolve_paper_file(conf_dir: Path) -> Path:
    """Resolve the paper JSON file for a workspace.

    Resolution order:
    1. workspace.json → "papers" path (explicit)
    2. papers.json in the workspace root (convention)
    3. Legacy fallback: first non-system JSON in root directory
    """
    ws_json = conf_dir / "workspace.json"
    if ws_json.is_file():
        try:
            with open(ws_json) as f:
                meta = json.load(f)
            if meta.get("papers"):
                explicit = conf_dir / meta["papers"]
                if explicit.is_file():
                    return explicit
        except Exception:
            pass

    # Convention: papers.json
    papers_file = conf_dir / "papers.json"
    if papers_file.is_file():
        return papers_file

    # Legacy fallback: first non-system JSON in root
    _SKIP_NAMES = {"metadata", "workspace", "token_usage", "oral_progress", "poster_progress"}
    for f in sorted(conf_dir.glob("*.json")):
        if not any(skip in f.name.lower() for skip in _SKIP_NAMES):
            return f

    raise ValueError(f"No paper data found in {conf_dir}")


def _parse_papers_json(paper_file: Path) -> list[Paper]:
    """Parse a paper JSON file into Paper objects."""
    with open(paper_file) as f:
        data = json.load(f)

    # Also look for a companion metadata file for abstracts
    abstract_lookup: dict[str, str] = {}
    meta_file = paper_file.parent / "metadata.json"
    if meta_file.is_file():
        try:
            with open(meta_file) as mf:
                meta = json.load(mf)
            for m in meta:
                if m.get("abstract"):
                    abstract_lookup[m["title"]] = m["abstract"]
        except Exception:
            pass

    papers = []
    for e in data:
        paper_id = str(e.get("id", e.get("paper_id", "")))
        raw_authors = e.get("authors", [])
        if isinstance(raw_authors, str):
            authors = [a.strip() for a in raw_authors.split(",") if a.strip()]
        else:
            authors = raw_authors
        abstract = e.get("abstract", "")
        if not abstract:
            abstract = abstract_lookup.get(e.get("title", ""), "")
        papers.append(Paper(id=paper_id, title=e["title"],
                            abstract=abstract, authors=authors))

    logger.info(f"Loaded {len(papers)} papers from {paper_file}")
    return papers


def load_conference_papers(conference: str) -> list[Paper]:
    """Load papers from a workspace directory."""
    conf_dir = DATA_DIR / conference
    if not conf_dir.is_dir():
        raise ValueError(f"Conference directory not found: {conf_dir}")
    paper_file = _resolve_paper_file(conf_dir)
    return _parse_papers_json(paper_file)


def get_workspace_mode(conference: str) -> str:
    """Read the mode (oral/poster) from workspace.json. Defaults to 'oral'."""
    ws_json = DATA_DIR / conference / "workspace.json"
    if ws_json.is_file():
        try:
            with open(ws_json) as f:
                return json.load(f).get("mode", "oral")
        except Exception:
            pass
    return "oral"


# ── Caches ──

_paper_cache: dict[str, list[Paper]] = {}
_taxonomy_cache: dict[str, TaxonomyNode] = {}
_similarity_cache: dict[str, SimilarityEngine] = {}


def get_papers(conference: str, mode: str = "oral") -> list[Paper]:
    if conference not in _paper_cache:
        _paper_cache[conference] = load_conference_papers(conference)
    return _paper_cache[conference]


def get_taxonomy(conference: str, papers: list[Paper]) -> TaxonomyNode:
    """Get or build taxonomy. Uses demo taxonomy if no LLM key is set."""
    if conference not in _taxonomy_cache:
        has_api_key = bool(
            os.environ.get("OPENROUTER_API_KEY")
            or os.environ.get("OPENAI_API_KEY")
            or os.environ.get("GOOGLE_API_KEY")
            or os.environ.get("GEMINI_API_KEY")
            or os.environ.get("ANTHROPIC_API_KEY")
            or os.environ.get("XAI_API_KEY")
        )

        if has_api_key:
            logger.info(f"Building LLM taxonomy for {conference}...")
            llm = LLMClient()
            builder = TaxonomyBuilder(papers, llm=llm,
                                      use_abstracts=getattr(config, "USE_ABSTRACTS", True))
            root = builder.build()
        else:
            logger.info(f"No LLM API key found; building automatic taxonomy for {conference}...")
            root = _build_auto_taxonomy(papers)

        _taxonomy_cache[conference] = root
    return _taxonomy_cache[conference]


def _build_auto_taxonomy(papers: list[Paper]) -> TaxonomyNode:
    """Build a simple automatic taxonomy by clustering papers using similarity."""
    # Simple single-level taxonomy: all papers in one group
    # The session organizer will handle the actual splitting
    root = TaxonomyNode(
        node_id="0",
        name="All Papers",
        description="Root node containing all accepted papers",
        paper_ids=[p.id for p in papers],
        depth=0,
        is_leaf=False,
    )

    # Try to create meaningful groups using TF-IDF clustering
    try:
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.cluster import KMeans
        import numpy as np

        texts = [p.text_for_embedding() for p in papers]
        vectorizer = TfidfVectorizer(max_features=3000, stop_words="english")
        tfidf = vectorizer.fit_transform(texts)

        n_clusters = min(max(3, len(papers) // 8), 15)
        kmeans = KMeans(n_clusters=n_clusters, random_state=42, n_init=10)
        labels = kmeans.fit_predict(tfidf)

        # Get top terms per cluster for naming
        feature_names = vectorizer.get_feature_names_out()
        cluster_groups: dict[int, list[str]] = defaultdict(list)
        for i, label in enumerate(labels):
            cluster_groups[label].append(papers[i].id)

        root.paper_ids = []
        root.is_leaf = False

        for idx, (cluster_id, paper_ids) in enumerate(sorted(cluster_groups.items())):
            # Name the cluster by top TF-IDF terms of its papers
            cluster_tfidf = tfidf[labels == cluster_id].toarray().mean(axis=0)
            top_term_indices = cluster_tfidf.argsort()[-3:][::-1]
            cluster_name = " & ".join(
                feature_names[i].title() for i in top_term_indices
            )

            child = TaxonomyNode(
                node_id=f"0.{idx}",
                name=cluster_name,
                description=f"Auto-clustered group of {len(paper_ids)} papers",
                parent_id="0",
                paper_ids=paper_ids,
                depth=1,
                is_leaf=True,
            )
            root.children.append(child)

        logger.info(f"Auto-taxonomy: {n_clusters} clusters for {len(papers)} papers")

    except Exception as e:
        logger.warning(f"Auto-taxonomy clustering failed, using flat: {e}")
        root.is_leaf = True

    return root


def get_similarity(conference: str, papers: list[Paper]) -> SimilarityEngine:
    if conference not in _similarity_cache:
        papers_map = {p.id: p for p in papers}
        engine = SimilarityEngine(
            papers_map,
            method=config.SIMILARITY_METHOD,
            use_cache=getattr(config, "EMBEDDING_CACHE_ENABLED", True),
        )
        engine.build()
        _similarity_cache[conference] = engine
    return _similarity_cache[conference]


def get_presenter_stats(papers: list[Paper]) -> dict:
    """Compute presenter / author statistics."""
    author_papers: dict[str, list[str]] = defaultdict(list)
    for p in papers:
        for a in p.authors:
            author_papers[a.strip()].append(p.id)

    multi = {a: pids for a, pids in author_papers.items() if len(pids) > 1}
    max_papers = max((len(pids) for pids in author_papers.values()), default=0)

    return {
        "presenterCount": len(author_papers),
        "multiPresenterCount": len(multi),
        "maxPapersPerPresenter": max_papers,
    }


# ════════════════════════════════════════════════════════════════════
# FastAPI app
# ════════════════════════════════════════════════════════════════════

app = FastAPI(title="TaxoConf API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── Serve index.html and static files ──

@app.get("/")
async def serve_index():
    index_path = PROJECT_ROOT / "index.html"
    if index_path.is_file():
        return FileResponse(str(index_path), media_type="text/html")
    return JSONResponse({"error": "index.html not found"}, status_code=404)


# Static file extensions allowed for catch-all route (registered at end)
STATIC_EXTENSIONS = {".css", ".js", ".png", ".jpg", ".svg", ".ico", ".woff", ".woff2", ".ttf"}


# ════════════════════════════════════════════════════════════════════
# Suggested capacity defaults from paper count
# ════════════════════════════════════════════════════════════════════

import math


def compute_oral_defaults(paper_count: int) -> dict:
    """Compute sensible oral session parameters from the number of papers."""
    if paper_count <= 0:
        return {"parallel_sessions": 1, "time_slots": 1,
                "max_per_session": 5, "min_per_session": 3}

    min_ps, max_ps = 3, 5
    avg_ps = (min_ps + max_ps) / 2  # 4
    total_sessions = math.ceil(paper_count / avg_ps)

    # Choose parallel tracks (M): scale with sqrt of sessions, clamp 2-8
    M = min(8, max(2, round(math.sqrt(total_sessions))))
    # Derive time slots (N)
    N = math.ceil(total_sessions / M)

    # Ensure capacity bounds are valid: M*N*min <= paper_count <= M*N*max
    # If too tight, widen by bumping N
    while M * N * min_ps > paper_count and N > 1:
        N -= 1
    while M * N * max_ps < paper_count:
        N += 1

    return {
        "parallel_sessions": M,
        "time_slots": N,
        "max_per_session": max_ps,
        "min_per_session": min_ps,
    }


def compute_poster_defaults(paper_count: int) -> dict:
    """Compute sensible poster session parameters from the number of papers."""
    if paper_count <= 0:
        return {"session_count": 1, "rows": 3, "cols": 4, "board_count": 12}

    rows, cols = 3, 4
    board_count = rows * cols  # 12 boards per session
    session_count = math.ceil(paper_count / board_count)

    return {
        "session_count": session_count,
        "rows": rows,
        "cols": cols,
        "board_count": board_count,
    }


# ════════════════════════════════════════════════════════════════════
# API: Oral Session Organization
# ════════════════════════════════════════════════════════════════════

@app.get("/api/oral/info")
async def oral_info(conference: str = Query("SIGIR25")):
    """Return info about available oral session data."""
    try:
        conferences = discover_conferences()
        # Normalize conference name
        conf = _resolve_conference(conference, conferences)

        papers = get_papers(conf)
        stats = get_presenter_stats(papers)
        suggested = compute_oral_defaults(len(papers))

        return {
            "result": {
                "conference": conf,
                "availableConferences": conferences,
                "paperCount": len(papers),
                "presenterCount": stats["presenterCount"],
                "multiPresenterCount": stats["multiPresenterCount"],
                "maxPapersPerPresenter": stats["maxPapersPerPresenter"],
                "paperDataPath": f"data/{conf}/",
                "similarityMatrixPath": f".cache/embeddings/ (auto-computed)",
                "suggested_params": suggested,
            }
        }
    except Exception as e:
        logger.error(f"oral/info error: {e}\n{traceback.format_exc()}")
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/api/oral/run")
async def oral_run(request: Request):
    """Run oral session organization."""
    try:
        body = await request.json()
        conference = body.get("conference", "SIGIR25")
        parallel_sessions = int(body.get("parallel_sessions", 7))
        time_slots = int(body.get("time_slots", 19))
        max_per_session = int(body.get("max_per_session", 4))
        min_per_session = int(body.get("min_per_session", 3))
        use_abstracts = body.get("use_abstracts", True)

        conferences = discover_conferences()
        conf = _resolve_conference(conference, conferences)

        # Override config for this run
        config.SESSION_MIN = min_per_session
        config.SESSION_MAX = max_per_session
        config.NUM_SLOTS = time_slots
        config.NUM_PARALLEL_TRACKS = parallel_sessions
        config.USE_ABSTRACTS = bool(use_abstracts)

        papers = get_papers(conf)
        papers_map = {p.id: p for p in papers}
        taxonomy_root = get_taxonomy(conf, papers)

        logger.info(f"Running oral organization: {len(papers)} papers, "
                     f"{parallel_sessions}x{time_slots} grid, "
                     f"session size {min_per_session}-{max_per_session}")

        result = run_oral_organization(papers, taxonomy_root)

        # Context-aware session naming (bottom-up cascade)
        try:
            naming_llm = LLMClient()
            name_sessions(result.sessions, taxonomy_root, papers_map, naming_llm)
        except Exception as e:
            logger.warning(f"Session naming failed, keeping original names: {e}")

        # Build response in the format the frontend expects.
        # The frontend renders a 2-D grid (slot × track) and looks up sessions
        # by id = "slot_{slot}_track_{track}" (both 1-indexed).  The backend
        # produces 0-indexed time_slot/track, so we convert here.
        sessions_out = []
        assignment_map = {}  # paper_id -> session_id (frontend format)

        for s in result.sessions:
            slot_1 = (s.time_slot or 0) + 1   # 0-indexed → 1-indexed
            track_1 = (s.track or 0) + 1
            frontend_id = f"slot_{slot_1}_track_{track_1}"

            session_data = {
                "id": frontend_id,
                "sessionName": s.name,
                "description": s.description,
                "slot": slot_1,
                "track": track_1,
                "paperCount": len(s.paper_ids),
                "targetSize": max_per_session,
                "papers": [],
            }
            for pid in s.paper_ids:
                p = papers_map.get(pid)
                if p:
                    session_data["papers"].append({
                        "id": p.id,
                        "title": p.title,
                        "abstract": p.abstract,
                        "authors": p.authors,
                        "presenters": p.authors,
                    })
                    assignment_map[p.id] = frontend_id

            sessions_out.append(session_data)

        # All papers list
        all_papers = []
        for p in papers:
            all_papers.append({
                "id": p.id,
                "title": p.title,
                "abstract": p.abstract,
                "authors": p.authors,
                "presenters": p.authors,
            })

        # ── LLM-based session review: flag misplaced papers ──
        hard_papers, review_status = _llm_review_sessions(sessions_out, mode="oral")

        response = {
            "result": {
                "sessions": sessions_out,
                "papers": all_papers,
                "assignment": assignment_map,
                "hardPapers": hard_papers,
                "reviewStatus": review_status,
                "stats": result.stats,
                "parallelSessions": parallel_sessions,
                "timeSlots": time_slots,
                "maxPerSession": max_per_session,
                "minPerSession": min_per_session,
                "tokenUsage": get_global_tracker().to_dict(),
            }
        }

        # Persist token stats
        try:
            tracker_data = get_global_tracker().to_dict()
            save_run_token_stats(conf, "oral", {
                "calls": tracker_data.get("total_calls", 0),
                "prompt_tokens": tracker_data.get("total_prompt_tokens", 0),
                "completion_tokens": tracker_data.get("total_completion_tokens", 0),
                "total_tokens": tracker_data.get("total_tokens", 0),
                "cost_usd": tracker_data.get("total_cost_usd", 0.0),
                "provider": config.LLM_PROVIDER,
                "model": config.LLM_MODEL,
            })
        except Exception as tok_err:
            logger.warning(f"Failed to save token stats: {tok_err}")

        return response

    except Exception as e:
        logger.error(f"oral/run error: {e}\n{traceback.format_exc()}")
        return JSONResponse({"error": str(e)}, status_code=500)


def _ensure_live_pricing():
    """Fetch live pricing from OpenRouter if not already cached."""
    from token_tracker import _live_pricing, set_live_pricing
    if _live_pricing:
        return  # Already loaded
    api_key = os.environ.get("OPENROUTER_API_KEY") or _manual_api_keys.get("openrouter")
    if not api_key:
        return
    try:
        import httpx
        headers = {"Authorization": f"Bearer {api_key}"}
        resp = httpx.get("https://openrouter.ai/api/v1/models", headers=headers, timeout=15)
        data = resp.json().get("data", [])
        pricing_list = []
        for m in data:
            pricing = m.get("pricing", {})
            prompt_price = float(pricing.get("prompt", "0") or "0")
            completion_price = float(pricing.get("completion", "0") or "0")
            if prompt_price > 0 or completion_price > 0:
                pricing_list.append({
                    "id": m.get("id", ""),
                    "prompt_price_per_1m": round(prompt_price * 1_000_000, 4),
                    "completion_price_per_1m": round(completion_price * 1_000_000, 4),
                })
        set_live_pricing(pricing_list)
    except Exception as e:
        logger.warning(f"Failed to fetch live pricing: {e}")


@app.post("/api/oral/run-stream")
async def oral_run_stream(request: Request):
    """Run oral session organization with SSE progress streaming."""
    from starlette.responses import StreamingResponse

    body = await request.json()

    def generate():
        try:
            conference = body.get("conference", "SIGIR25")
            parallel_sessions = int(body.get("parallel_sessions", 7))
            time_slots = int(body.get("time_slots", 19))
            max_per_session = int(body.get("max_per_session", 4))
            min_per_session = int(body.get("min_per_session", 3))
            use_abstracts = body.get("use_abstracts", True)

            conferences = discover_conferences()
            conf = _resolve_conference(conference, conferences)

            config.SESSION_MIN = min_per_session
            config.SESSION_MAX = max_per_session
            config.NUM_SLOTS = time_slots
            config.NUM_PARALLEL_TRACKS = parallel_sessions
            config.USE_ABSTRACTS = bool(use_abstracts)

            # Ensure live pricing is loaded for accurate cost tracking
            _ensure_live_pricing()

            # Step 1: Load papers + build similarity
            yield f"data: {json.dumps({'type':'progress','step':1,'total':7,'msg':'Loading papers and building similarity matrix...'})}\n\n"
            papers = get_papers(conf)
            papers_map = {p.id: p for p in papers}

            # Step 2: Taxonomy construction (clear cache to force rebuild)
            _taxonomy_cache.pop(conf, None)
            yield f"data: {json.dumps({'type':'progress','step':2,'total':7,'msg':f'Constructing topic taxonomy for {len(papers)} papers via LLM...'})}\n\n"
            taxonomy_root = get_taxonomy(conf, papers)

            # Step 3: Session formation
            yield f"data: {json.dumps({'type':'progress','step':3,'total':7,'msg':'Forming sessions from taxonomy leaves...'})}\n\n"
            from session_organizer import run_oral_organization as _run_oral
            result = _run_oral(papers, taxonomy_root)

            # Step 4: Session naming (bottom-up cascade)
            yield f"data: {json.dumps({'type':'progress','step':4,'total':7,'msg':f'Generating names for {len(result.sessions)} sessions (bottom-up cascade)...'})}\n\n"
            try:
                naming_llm = LLMClient()
                # name_sessions includes normalization
                from session_namer import name_sessions as _name
                _name(result.sessions, taxonomy_root, papers_map, naming_llm)
            except Exception as e:
                logger.warning(f"Session naming failed: {e}")

            # Step 5: Build response
            yield f"data: {json.dumps({'type':'progress','step':5,'total':7,'msg':'Building schedule grid...'})}\n\n"
            sessions_out = []
            assignment_map = {}
            for s in result.sessions:
                slot_1 = (s.time_slot or 0) + 1
                track_1 = (s.track or 0) + 1
                frontend_id = f"slot_{slot_1}_track_{track_1}"
                session_data = {
                    "id": frontend_id,
                    "sessionName": s.name,
                    "description": s.description,
                    "slot": slot_1, "track": track_1,
                    "paperCount": len(s.paper_ids),
                    "targetSize": max_per_session,
                    "papers": [],
                }
                for pid in s.paper_ids:
                    p = papers_map.get(pid)
                    if p:
                        session_data["papers"].append({
                            "id": p.id, "title": p.title,
                            "abstract": p.abstract, "authors": p.authors,
                            "presenters": p.authors,
                        })
                        assignment_map[p.id] = frontend_id
                sessions_out.append(session_data)

            all_papers = [{"id": p.id, "title": p.title, "abstract": p.abstract,
                           "authors": p.authors, "presenters": p.authors} for p in papers]

            # Step 6: LLM session review
            yield f"data: {json.dumps({'type':'progress','step':6,'total':7,'msg':'Reviewing sessions for misplaced papers...'})}\n\n"
            hard_papers, review_status = _llm_review_sessions(sessions_out, mode="oral")

            # Step 7: Finalize
            yield f"data: {json.dumps({'type':'progress','step':7,'total':7,'msg':'Finalizing results...'})}\n\n"

            # Save token stats
            try:
                tracker_data = get_global_tracker().to_dict()
                save_run_token_stats(conf, "oral", {
                    "calls": tracker_data.get("total_calls", 0),
                    "prompt_tokens": tracker_data.get("total_prompt_tokens", 0),
                    "completion_tokens": tracker_data.get("total_completion_tokens", 0),
                    "total_tokens": tracker_data.get("total_tokens", 0),
                    "cost_usd": tracker_data.get("total_cost_usd", 0.0),
                    "provider": config.LLM_PROVIDER,
                    "model": config.LLM_MODEL,
                })
            except Exception:
                pass

            response = {
                "result": {
                    "sessions": sessions_out,
                    "papers": all_papers,
                    "assignment": assignment_map,
                    "hardPapers": hard_papers,
                    "reviewStatus": review_status,
                    "stats": result.stats,
                    "parallelSessions": parallel_sessions,
                    "timeSlots": time_slots,
                    "maxPerSession": max_per_session,
                    "minPerSession": min_per_session,
                    "tokenUsage": get_global_tracker().to_dict(),
                }
            }
            yield f"data: {json.dumps({'type':'result','data':response})}\n\n"

        except Exception as e:
            logger.error(f"oral/run-stream error: {e}\n{traceback.format_exc()}")
            yield f"data: {json.dumps({'type':'error','error':str(e)})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream")


# ── Excel export ─────────────────────────────────────────────────────

@app.post("/api/{mode}/export-excel")
async def export_excel(mode: str, request: Request):
    """Export session schedule to Excel using the template."""
    if mode not in ("oral", "poster"):
        return JSONResponse({"error": "Invalid mode"}, status_code=400)
    try:
        import openpyxl
        from io import BytesIO

        body = await request.json()
        sessions = body.get("sessions", [])
        track_names = body.get("trackNames", [])

        template_path = PROJECT_ROOT / "template" / "excel_template.xlsx"
        if not template_path.is_file():
            return JSONResponse({"error": "Excel template not found"}, status_code=404)

        wb = openpyxl.load_workbook(str(template_path))
        ws = wb["Agenda"]

        row = 25  # Start data at row 25
        for session in sessions:
            s_date = session.get("sessionDate", "")
            s_start = session.get("startTime", "")
            s_end = session.get("endTime", "")
            s_track_idx = session.get("track", 0)
            s_track_name = ""
            if track_names and 0 < s_track_idx <= len(track_names):
                s_track_name = track_names[s_track_idx - 1]
            elif session.get("trackLabel"):
                s_track_name = session["trackLabel"]
            s_title = session.get("sessionName", "")
            s_room = session.get("location", "")
            s_chair = session.get("sessionChair", "")

            # Write session row
            ws.cell(row=row, column=1, value=s_date)       # Date
            ws.cell(row=row, column=2, value=s_start)      # Time Start
            ws.cell(row=row, column=3, value=s_end)        # Time End
            ws.cell(row=row, column=4, value=s_track_name) # Tracks
            ws.cell(row=row, column=5, value=s_title)      # Session Title
            ws.cell(row=row, column=6, value=s_room)       # Room/Location
            ws.cell(row=row, column=7, value="")           # Description (empty)
            ws.cell(row=row, column=8, value=s_chair)      # Speakers (chair)
            ws.cell(row=row, column=9, value="")           # Authors (empty)
            ws.cell(row=row, column=10, value="")          # Session/Sub (empty)
            row += 1

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        from starlette.responses import Response
        filename = f"{mode}_schedule.xlsx"
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except ImportError:
        return JSONResponse(
            {"error": "openpyxl is required for Excel export. Install with: pip install openpyxl"},
            status_code=500,
        )
    except Exception as e:
        logger.error(f"{mode}/export-excel error: {e}\n{traceback.format_exc()}")
        return JSONResponse({"error": str(e)}, status_code=500)


# ── Progress save/load (oral + poster, with named saves) ────────────

import re as _re


def _safe_save_name(name: str) -> str:
    """Sanitize a save name for use as a filename component."""
    name = _re.sub(r'[^\w\-]', '_', name.strip())[:60]
    return name or "default"


def _progress_dir(conf: str, mode: str) -> Path:
    d = PROJECT_ROOT / "data" / conf / f"{mode}_progress"
    d.mkdir(parents=True, exist_ok=True)
    return d


@app.get("/api/{mode}/progress/list")
async def list_progress_saves(mode: str, conference: str = Query("SIGIR25")):
    """List all named saves for a mode (oral/poster)."""
    if mode not in ("oral", "poster"):
        return JSONResponse({"error": "Invalid mode"}, status_code=400)
    try:
        conferences = discover_conferences()
        conf = _resolve_conference(conference, conferences)
        d = _progress_dir(conf, mode)
        saves = []
        for f in sorted(d.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True):
            saves.append({
                "name": f.stem,
                "modified": f.stat().st_mtime,
                "size": f.stat().st_size,
            })
        # Also check legacy single-file progress
        legacy = PROJECT_ROOT / "data" / conf / f"{mode}_progress.json"
        if legacy.is_file() and not (d / "default.json").is_file():
            saves.insert(0, {
                "name": "default (legacy)",
                "modified": legacy.stat().st_mtime,
                "size": legacy.stat().st_size,
            })
        return {"success": True, "saves": saves}
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.post("/api/{mode}/progress")
async def save_progress(mode: str, request: Request):
    """Save progress with an optional name."""
    if mode not in ("oral", "poster"):
        return JSONResponse({"error": "Invalid mode"}, status_code=400)
    try:
        body = await request.json()
        conference = body.get("conference", "SIGIR25")
        result_data = body.get("result")
        save_name = _safe_save_name(body.get("name", "default"))
        if not result_data:
            return {"success": False, "error": "No result data provided"}
        conferences = discover_conferences()
        conf = _resolve_conference(conference, conferences)
        d = _progress_dir(conf, mode)
        path = d / f"{save_name}.json"
        import json
        with open(path, "w", encoding="utf-8") as f:
            json.dump(result_data, f, ensure_ascii=False, indent=2)
        logger.info(f"{mode} progress saved: {path}")
        return {"success": True, "path": str(path), "name": save_name}
    except Exception as e:
        logger.error(f"{mode}/progress save error: {e}\n{traceback.format_exc()}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.get("/api/{mode}/progress")
async def load_progress(mode: str, conference: str = Query("SIGIR25"),
                        name: str = Query("default")):
    """Load a named progress save."""
    if mode not in ("oral", "poster"):
        return JSONResponse({"error": "Invalid mode"}, status_code=400)
    try:
        conferences = discover_conferences()
        conf = _resolve_conference(conference, conferences)
        save_name = _safe_save_name(name)
        d = _progress_dir(conf, mode)
        path = d / f"{save_name}.json"
        # Fall back to legacy single-file if named file doesn't exist
        if not path.is_file():
            legacy = PROJECT_ROOT / "data" / conf / f"{mode}_progress.json"
            if legacy.is_file():
                path = legacy
            else:
                return {"success": False, "error": f"No save named '{save_name}' found"}
        import json
        with open(path, "r", encoding="utf-8") as f:
            result_data = json.load(f)
        return {"success": True, "result": result_data, "name": save_name}
    except Exception as e:
        logger.error(f"{mode}/progress load error: {e}\n{traceback.format_exc()}")
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


# ════════════════════════════════════════════════════════════════════
# API: Poster Session Organization
# ════════════════════════════════════════════════════════════════════

@app.get("/api/poster/info")
async def poster_info(conference: str = Query("SIGIR25")):
    """Return info about available poster session data."""
    try:
        conferences = discover_conferences()
        conf = _resolve_conference(conference, conferences)

        papers = get_papers(conf)
        stats = get_presenter_stats(papers)
        suggested = compute_poster_defaults(len(papers))

        return {
            "result": {
                "conference": conf,
                "availableConferences": conferences,
                "paperCount": len(papers),
                "presenterCount": stats["presenterCount"],
                "multiPresenterCount": stats["multiPresenterCount"],
                "maxPapersPerPresenter": stats["maxPapersPerPresenter"],
                "paperDataPath": f"data/{conf}/",
                "similarityMatrixPath": f".cache/embeddings/ (auto-computed)",
                "suggested_params": suggested,
            }
        }
    except Exception as e:
        logger.error(f"poster/info error: {e}\n{traceback.format_exc()}")
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/api/poster/run")
async def poster_run(request: Request):
    """Run poster session organization."""
    try:
        body = await request.json()
        conference = body.get("conference", "SIGIR25")
        layout_type = body.get("layout_type", "rectangle")
        board_count = int(body.get("board_count", 12))
        rows = int(body.get("rows", 3))
        cols = int(body.get("cols", 4))
        session_count = int(body.get("session_count", 44))
        prevent_same_presenter = bool(body.get("prevent_same_presenter", False))
        optimize_within = bool(body.get("optimize_within_layout", True))
        use_abstracts = body.get("use_abstracts", True)

        conferences = discover_conferences()
        conf = _resolve_conference(conference, conferences)

        # Compute session capacity
        if layout_type == "rectangle":
            session_capacity = rows * cols
        else:
            session_capacity = board_count

        # Override config for this run
        config.POSTER_SESSION_MIN = max(1, session_capacity // 2)
        config.POSTER_SESSION_MAX = session_capacity
        config.POSTER_NUM_SLOTS = max(1, session_count // max(1, 2))  # rough: 2 parallel areas
        config.POSTER_NUM_PARALLEL = min(session_count, 2)
        config.POSTER_FLOOR_PLAN = layout_type
        config.POSTER_RECT_COLS = cols
        config.POSTER_PROXIMITY = optimize_within
        config.USE_ABSTRACTS = bool(use_abstracts)
        config.POSTER_ENABLE_CONFLICT_AVOIDANCE = prevent_same_presenter

        papers = get_papers(conf)
        papers_map = {p.id: p for p in papers}
        taxonomy_root = get_taxonomy(conf, papers)

        floor_plan = FloorPlanType(layout_type) if layout_type in ("line", "circle", "rectangle") else FloorPlanType.RECTANGLE

        logger.info(f"Running poster organization: {len(papers)} papers, "
                     f"{session_count} sessions, layout={layout_type}, "
                     f"capacity={session_capacity}")

        poster_result = run_poster_pipeline(
            papers=papers,
            taxonomy_root=taxonomy_root,
            floor_plan=floor_plan,
            rect_cols=cols,
            enable_proximity=optimize_within,
            avoid_conflicts=prevent_same_presenter,
            num_slots=config.POSTER_NUM_SLOTS,
            num_parallel=config.POSTER_NUM_PARALLEL,
        )

        # Context-aware session naming (bottom-up cascade)
        if poster_result.org_result:
            try:
                naming_llm = LLMClient()
                name_sessions(poster_result.org_result.sessions, taxonomy_root,
                              papers_map, naming_llm)
                org_session_names = {s.session_id: s.name for s in poster_result.org_result.sessions}
                for ps in poster_result.poster_sessions:
                    if ps.session_id in org_session_names:
                        ps.name = org_session_names[ps.session_id]
            except Exception as e:
                logger.warning(f"Poster session naming failed, keeping original names: {e}")

        # Build response
        sessions_out = []
        placements = {}  # paper_id -> {sessionId, cellIndex}

        for ps in poster_result.poster_sessions:
            session_data = {
                "id": ps.session_id,
                "sessionName": ps.name,
                "description": ps.description,
                "slot": ps.time_slot,
                "area": ps.area,
                "paperCount": len(ps.assignments),
                "cells": [],
                "papers": [],
            }

            # Build the cells array (fixed size = session_capacity)
            cell_map: dict[int, dict] = {}
            for a in ps.assignments:
                p = papers_map.get(a.paper_id)
                if p:
                    paper_data = {
                        "id": p.id,
                        "title": p.title,
                        "abstract": p.abstract,
                        "authors": p.authors,
                        "presenters": p.authors,
                    }
                    cell_map[a.board.index] = paper_data
                    session_data["papers"].append(paper_data)
                    placements[p.id] = {
                        "sessionId": ps.session_id,
                        "cellIndex": a.board.index,
                    }

            # Fill cells array
            for idx in range(session_capacity):
                session_data["cells"].append(cell_map.get(idx, None))

            sessions_out.append(session_data)

        # All papers
        all_papers = []
        for p in papers:
            all_papers.append({
                "id": p.id,
                "title": p.title,
                "abstract": p.abstract,
                "authors": p.authors,
                "presenters": p.authors,
            })

        # ── LLM-based session review: flag misplaced papers ──
        hard_papers, review_status = _llm_review_sessions(sessions_out, mode="poster")

        response = {
            "result": {
                "sessions": sessions_out,
                "papers": all_papers,
                "placements": placements,
                "hardPapers": hard_papers,
                "reviewStatus": review_status,
                "layoutType": layout_type,
                "rows": rows,
                "cols": cols,
                "boardCount": board_count,
                "sessionCount": len(sessions_out),
                "sessionCapacity": session_capacity,
                "optimizeWithinLayout": optimize_within,
                "stats": poster_result.stats,
                "tokenUsage": get_global_tracker().to_dict(),
            }
        }

        # Persist token stats
        try:
            tracker_data = get_global_tracker().to_dict()
            save_run_token_stats(conf, "poster", {
                "calls": tracker_data.get("total_calls", 0),
                "prompt_tokens": tracker_data.get("total_prompt_tokens", 0),
                "completion_tokens": tracker_data.get("total_completion_tokens", 0),
                "total_tokens": tracker_data.get("total_tokens", 0),
                "cost_usd": tracker_data.get("total_cost_usd", 0.0),
                "provider": config.LLM_PROVIDER,
                "model": config.LLM_MODEL,
            })
        except Exception as tok_err:
            logger.warning(f"Failed to save token stats: {tok_err}")

        return response

    except Exception as e:
        logger.error(f"poster/run error: {e}\n{traceback.format_exc()}")
        return JSONResponse({"error": str(e)}, status_code=500)


@app.post("/api/poster/run-stream")
async def poster_run_stream(request: Request):
    """Run poster session organization with SSE progress streaming."""
    from starlette.responses import StreamingResponse

    body = await request.json()

    def generate():
        try:
            conference = body.get("conference", "SIGIR25")
            layout_type = body.get("layout_type", "rectangle")
            board_count = int(body.get("board_count", 12))
            rows = int(body.get("rows", 3))
            cols = int(body.get("cols", 4))
            session_count = int(body.get("session_count", 44))
            prevent_same_presenter = bool(body.get("prevent_same_presenter", False))
            optimize_within = bool(body.get("optimize_within_layout", True))
            use_abstracts = body.get("use_abstracts", True)

            conferences = discover_conferences()
            conf = _resolve_conference(conference, conferences)

            if layout_type == "rectangle":
                session_capacity = rows * cols
            else:
                session_capacity = board_count

            config.POSTER_SESSION_MIN = max(1, session_capacity // 2)
            config.POSTER_SESSION_MAX = session_capacity
            config.POSTER_NUM_SLOTS = max(1, session_count // max(1, 2))
            config.POSTER_NUM_PARALLEL = min(session_count, 2)
            config.POSTER_FLOOR_PLAN = layout_type
            config.POSTER_RECT_COLS = cols
            config.POSTER_PROXIMITY = optimize_within
            config.USE_ABSTRACTS = bool(use_abstracts)
            config.POSTER_ENABLE_CONFLICT_AVOIDANCE = prevent_same_presenter

            _ensure_live_pricing()

            # Step 1
            yield f"data: {json.dumps({'type':'progress','step':1,'total':8,'msg':f'Loading {conference} papers and building similarity matrix...'})}\n\n"
            papers = get_papers(conf)
            papers_map = {p.id: p for p in papers}

            # Step 2 (clear cache to force rebuild)
            _taxonomy_cache.pop(conf, None)
            yield f"data: {json.dumps({'type':'progress','step':2,'total':8,'msg':f'Constructing topic taxonomy for {len(papers)} papers via LLM...'})}\n\n"
            taxonomy_root = get_taxonomy(conf, papers)

            floor_plan = FloorPlanType(layout_type) if layout_type in ("line", "circle", "rectangle") else FloorPlanType.RECTANGLE

            # Step 3
            yield f"data: {json.dumps({'type':'progress','step':3,'total':8,'msg':'Forming poster sessions and scheduling into time slots...'})}\n\n"
            poster_result = run_poster_pipeline(
                papers=papers, taxonomy_root=taxonomy_root,
                floor_plan=floor_plan, rect_cols=cols,
                enable_proximity=optimize_within,
                avoid_conflicts=prevent_same_presenter,
                num_slots=config.POSTER_NUM_SLOTS,
                num_parallel=config.POSTER_NUM_PARALLEL,
            )

            # Step 4
            yield f"data: {json.dumps({'type':'progress','step':4,'total':8,'msg':'Optimizing board layout for topical proximity...'})}\n\n"
            # (already done inside run_poster_pipeline, but status update is useful)

            # Step 5
            yield f"data: {json.dumps({'type':'progress','step':5,'total':8,'msg':f'Generating names for {len(poster_result.poster_sessions)} sessions...'})}\n\n"
            if poster_result.org_result:
                try:
                    naming_llm = LLMClient()
                    from session_namer import name_sessions as _name
                    _name(poster_result.org_result.sessions, taxonomy_root, papers_map, naming_llm)
                    org_names = {s.session_id: s.name for s in poster_result.org_result.sessions}
                    for ps in poster_result.poster_sessions:
                        if ps.session_id in org_names:
                            ps.name = org_names[ps.session_id]
                except Exception as e:
                    logger.warning(f"Poster naming failed: {e}")

            # Step 6
            yield f"data: {json.dumps({'type':'progress','step':6,'total':8,'msg':'Building poster grid layout...'})}\n\n"
            sessions_out = []
            placements = {}
            for ps in poster_result.poster_sessions:
                session_data = {
                    "id": ps.session_id, "sessionName": ps.name,
                    "description": ps.description,
                    "slot": ps.time_slot, "area": ps.area,
                    "paperCount": len(ps.assignments),
                    "cells": [], "papers": [],
                }
                cell_map = {}
                for a in ps.assignments:
                    p = papers_map.get(a.paper_id)
                    if p:
                        pd = {"id": p.id, "title": p.title, "abstract": p.abstract,
                              "authors": p.authors, "presenters": p.authors}
                        cell_map[a.board.index] = pd
                        session_data["papers"].append(pd)
                        placements[p.id] = {"sessionId": ps.session_id, "cellIndex": a.board.index}
                for idx in range(session_capacity):
                    session_data["cells"].append(cell_map.get(idx, None))
                sessions_out.append(session_data)

            all_papers = [{"id": p.id, "title": p.title, "abstract": p.abstract,
                           "authors": p.authors, "presenters": p.authors} for p in papers]

            # Step 7
            yield f"data: {json.dumps({'type':'progress','step':7,'total':8,'msg':'Reviewing sessions for misplaced papers...'})}\n\n"
            hard_papers, review_status = _llm_review_sessions(sessions_out, mode="poster")

            # Step 8
            yield f"data: {json.dumps({'type':'progress','step':8,'total':8,'msg':'Finalizing results...'})}\n\n"
            try:
                tracker_data = get_global_tracker().to_dict()
                save_run_token_stats(conf, "poster", {
                    "calls": tracker_data.get("total_calls", 0),
                    "prompt_tokens": tracker_data.get("total_prompt_tokens", 0),
                    "completion_tokens": tracker_data.get("total_completion_tokens", 0),
                    "total_tokens": tracker_data.get("total_tokens", 0),
                    "cost_usd": tracker_data.get("total_cost_usd", 0.0),
                    "provider": config.LLM_PROVIDER, "model": config.LLM_MODEL,
                })
            except Exception:
                pass

            response = {
                "result": {
                    "sessions": sessions_out, "papers": all_papers,
                    "placements": placements, "hardPapers": hard_papers,
                    "reviewStatus": review_status, "layoutType": layout_type,
                    "rows": rows, "cols": cols, "boardCount": board_count,
                    "sessionCount": len(sessions_out),
                    "sessionCapacity": session_capacity,
                    "optimizeWithinLayout": optimize_within,
                    "stats": poster_result.stats,
                    "tokenUsage": get_global_tracker().to_dict(),
                }
            }
            yield f"data: {json.dumps({'type':'result','data':response})}\n\n"

        except Exception as e:
            logger.error(f"poster/run-stream error: {e}\n{traceback.format_exc()}")
            yield f"data: {json.dumps({'type':'error','error':str(e)})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream")



# ════════════════════════════════════════════════════════════════════
# API: Paper Assignment (placeholder)
# ════════════════════════════════════════════════════════════════════

@app.get("/api/assignment/info")
async def assignment_info(conference: str = Query("SIGIR25")):
    """Placeholder for assignment info endpoint."""
    conferences = discover_conferences()
    conf = _resolve_conference(conference, conferences)
    try:
        papers = get_papers(conf)
        paper_count = len(papers)
    except Exception:
        paper_count = 0

    return {
        "result": {
            "conference": conf,
            "availableConferences": conferences,
            "paperCount": paper_count,
            "reviewerCount": 0,
            "metaReviewerCount": 0,
            "status": "placeholder",
            "message": "Paper assignment backend is not yet integrated. "
                       "This is a placeholder endpoint.",
        }
    }


@app.post("/api/assignment/run")
async def assignment_run(request: Request):
    """Placeholder for assignment run endpoint."""
    return JSONResponse(
        {"error": "Paper assignment backend is not yet integrated. "
                  "This endpoint is a placeholder."},
        status_code=501,
    )


# ════════════════════════════════════════════════════════════════════
# API: Reviewer Discovery (placeholder)
# ════════════════════════════════════════════════════════════════════

@app.get("/api/discovery/info")
async def discovery_info():
    """Placeholder for reviewer discovery."""
    return {
        "result": {
            "status": "in_progress",
            "message": "Reviewer discovery is still being integrated.",
        }
    }


# ════════════════════════════════════════════════════════════════════
# Helpers
# ════════════════════════════════════════════════════════════════════

def _resolve_conference(name: str, available: list[str]) -> str:
    """Resolve a conference name to one of the available directories.

    Tries exact match, then case-insensitive, then partial match.
    """
    if not available:
        raise ValueError("No conference data directories found under ./data/")

    # Exact match
    if name in available:
        return name

    # Case-insensitive
    lower_map = {c.lower(): c for c in available}
    if name.lower() in lower_map:
        return lower_map[name.lower()]

    # Partial match (e.g. "sigir2025" matches "SIGIR25")
    name_norm = name.lower().replace("-", "").replace("_", "").replace(" ", "")
    for c in available:
        c_norm = c.lower().replace("-", "").replace("_", "").replace(" ", "")
        if name_norm in c_norm or c_norm in name_norm:
            return c

    # Default to first available
    logger.warning(f"Conference '{name}' not found, defaulting to '{available[0]}'")
    return available[0]


def _llm_review_sessions(sessions_out: list[dict], mode: str = "oral"
                         ) -> tuple[list[dict], dict]:
    """Run LLM-based review on organized sessions to flag misplaced papers.

    Returns (hard_papers, review_status) where review_status contains
    diagnostic info for the API response.
    """
    has_api_key = bool(
        os.environ.get("OPENROUTER_API_KEY")
        or os.environ.get("OPENAI_API_KEY")
        or os.environ.get("GOOGLE_API_KEY")
        or os.environ.get("GEMINI_API_KEY")
        or os.environ.get("ANTHROPIC_API_KEY")
        or os.environ.get("XAI_API_KEY")
    )
    if not has_api_key:
        logger.info(f"No LLM API key found; skipping {mode} session review")
        return [], {"status": "skipped", "reason": "No LLM API key configured"}

    try:
        llm = LLMClient()
        logger.info(f"LLM session review ({mode}): using provider={llm.provider}, "
                     f"model={llm.model}")
        hard_papers = review_sessions(
            llm,
            sessions_out,
            all_sessions=sessions_out,
            mode=mode,
        )
        return hard_papers, {
            "status": "completed",
            "provider": llm.provider,
            "model": llm.model,
            "sessionsReviewed": len(sessions_out),
            "flaggedCount": len(hard_papers),
        }
    except Exception as e:
        error_msg = f"{type(e).__name__}: {e}"
        logger.error(f"LLM session review ({mode}) failed: {error_msg}")
        import traceback
        logger.error(traceback.format_exc())
        return [], {"status": "error", "reason": error_msg}


# ════════════════════════════════════════════════════════════════════
# API: Workspace Management
# ════════════════════════════════════════════════════════════════════

def _ensure_workspace_json(ws_dir: Path, name: str):
    """Create workspace.json if it doesn't exist (migration for legacy dirs)."""
    ws_json = ws_dir / "workspace.json"
    if ws_json.exists():
        return
    # Auto-migrate: look for existing paper JSON to get counts
    paper_count = 0
    for f in ws_dir.glob("*.json"):
        if "metadata" not in f.name.lower() and "workspace" not in f.name.lower():
            try:
                with open(f) as fh:
                    data = json.load(fh)
                    paper_count = len(data) if isinstance(data, list) else 0
            except Exception:
                pass
            break

    import datetime
    meta = {
        "name": name,
        "created_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "description": f"Auto-migrated workspace for {name}",
        "paper_count": paper_count,
    }
    with open(ws_json, "w") as fh:
        json.dump(meta, fh, indent=2)
    logger.info(f"Created workspace.json for '{name}' (migrated)")


def _list_workspaces() -> list[dict]:
    """List all workspaces under DATA_DIR."""
    workspaces = []
    if not DATA_DIR.is_dir():
        return workspaces
    for entry in sorted(DATA_DIR.iterdir()):
        if entry.is_dir() and not entry.name.startswith("."):
            _ensure_workspace_json(entry, entry.name)
            ws_json = entry / "workspace.json"
            if ws_json.exists():
                try:
                    with open(ws_json) as fh:
                        meta = json.load(fh)
                    meta["name"] = entry.name  # ensure name matches dir
                    workspaces.append(meta)
                except Exception:
                    workspaces.append({"name": entry.name})
    return workspaces


@app.get("/api/workspaces")
async def list_workspaces():
    """List all available workspaces."""
    return {"result": _list_workspaces()}


@app.post("/api/workspaces")
async def create_workspace(request: Request):
    """Create a new workspace directory."""
    body = await request.json()
    name = body.get("name", "").strip()
    if not name:
        return JSONResponse({"error": "Workspace name is required."}, status_code=400)
    # Sanitize name
    safe_name = "".join(c for c in name if c.isalnum() or c in "-_ ").strip()
    if not safe_name:
        return JSONResponse({"error": "Invalid workspace name."}, status_code=400)

    ws_dir = DATA_DIR / safe_name
    if ws_dir.exists():
        return JSONResponse({"error": f"Workspace '{safe_name}' already exists."}, status_code=409)

    ws_dir.mkdir(parents=True, exist_ok=True)

    import datetime
    mode = body.get("mode", "oral")
    if mode not in ("oral", "poster"):
        mode = "oral"

    meta = {
        "name": safe_name,
        "created_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "description": body.get("description", ""),
        "mode": mode,
        "paper_count": 0,
    }
    with open(ws_dir / "workspace.json", "w") as fh:
        json.dump(meta, fh, indent=2)

    logger.info(f"Created workspace '{safe_name}'")
    return {"success": True, "workspace": meta}


@app.get("/api/workspaces/{name}")
async def get_workspace(name: str):
    """Get workspace details."""
    ws_dir = DATA_DIR / name
    if not ws_dir.is_dir():
        return JSONResponse({"error": f"Workspace '{name}' not found."}, status_code=404)

    _ensure_workspace_json(ws_dir, name)
    ws_json = ws_dir / "workspace.json"
    try:
        with open(ws_json) as fh:
            meta = json.load(fh)
    except Exception:
        meta = {"name": name}

    # Count papers from any JSON file
    paper_count = 0
    for f in ws_dir.glob("*.json"):
        if "metadata" not in f.name.lower() and "workspace" not in f.name.lower() and "token" not in f.name.lower():
            try:
                with open(f) as fh:
                    data = json.load(fh)
                    paper_count = len(data) if isinstance(data, list) else 0
            except Exception:
                pass
            break
    meta["paper_count"] = paper_count
    return {"result": meta}


@app.post("/api/workspaces/{name}/upload")
async def upload_workspace_papers(name: str, request: Request):
    """Upload a paper JSON file to a workspace. Saved as papers.json."""
    ws_dir = DATA_DIR / name
    if not ws_dir.is_dir():
        return JSONResponse({"error": f"Workspace '{name}' not found."}, status_code=404)

    body = await request.json()
    papers = body.get("papers", [])

    if not papers or not isinstance(papers, list):
        return JSONResponse({"error": "Invalid papers data. Expected a JSON array."}, status_code=400)

    # Save as papers.json in the workspace root
    out_path = ws_dir / "papers.json"
    with open(out_path, "w") as fh:
        json.dump(papers, fh, indent=2)

    # Update workspace.json
    _ensure_workspace_json(ws_dir, name)
    ws_json = ws_dir / "workspace.json"
    try:
        with open(ws_json) as fh:
            meta = json.load(fh)
    except Exception:
        meta = {"name": name}
    meta["papers"] = "papers.json"
    meta["paper_count"] = len(papers)
    with open(ws_json, "w") as fh:
        json.dump(meta, fh, indent=2)

    # Clear caches for this mode
    _paper_cache.pop(f"{name}_oral", None)
    _paper_cache.pop(f"{name}_poster", None)
    _taxonomy_cache.pop(name, None)
    _similarity_cache.pop(name, None)

    logger.info(f"Uploaded {len(papers)} papers to workspace '{name}'")
    return {"success": True, "paper_count": len(papers)}


@app.delete("/api/workspaces/{name}")
async def delete_workspace(name: str):
    """Delete a workspace directory."""
    ws_dir = DATA_DIR / name
    if not ws_dir.is_dir():
        return JSONResponse({"error": f"Workspace '{name}' not found."}, status_code=404)

    import shutil
    try:
        shutil.rmtree(ws_dir)
    except PermissionError as pe:
        logger.warning(f"Permission error deleting workspace '{name}': {pe}")
        return JSONResponse(
            {"error": f"Cannot delete workspace '{name}': permission denied. "
                      "Please delete the folder manually."},
            status_code=403,
        )

    # Clear caches
    _paper_cache.pop(name, None)
    _taxonomy_cache.pop(name, None)
    _similarity_cache.pop(name, None)

    logger.info(f"Deleted workspace '{name}'")
    return {"success": True, "message": f"Workspace '{name}' deleted."}


# ════════════════════════════════════════════════════════════════════
# API: Token & Cost Statistics
# ════════════════════════════════════════════════════════════════════

_GLOBAL_TOKEN_FILE = DATA_DIR / "global_token_usage.json"


def _load_json_file(path: Path) -> dict:
    if path.exists():
        try:
            with open(path) as fh:
                return json.load(fh)
        except Exception:
            pass
    return {}


def _save_json_file(path: Path, data: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as fh:
        json.dump(data, fh, indent=2)


def _get_workspace_token_file(workspace: str) -> Path:
    return DATA_DIR / workspace / "token_usage.json"


def save_run_token_stats(workspace: str, mode: str, run_stats: dict):
    """Save token usage for a completed run to workspace and global files."""
    import datetime
    run_entry = {
        "run_id": f"{mode}_{datetime.datetime.now(datetime.timezone.utc).strftime('%Y%m%dT%H%M%S')}",
        "mode": mode,
        "timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        **run_stats,
    }

    # Update workspace token file
    ws_file = _get_workspace_token_file(workspace)
    ws_data = _load_json_file(ws_file)
    if "runs" not in ws_data:
        ws_data = {
            "workspace": workspace,
            "total_calls": 0, "total_prompt_tokens": 0,
            "total_completion_tokens": 0, "total_tokens": 0,
            "total_cost_usd": 0.0, "runs": [],
        }
    ws_data["total_calls"] += run_stats.get("calls", 0)
    ws_data["total_prompt_tokens"] += run_stats.get("prompt_tokens", 0)
    ws_data["total_completion_tokens"] += run_stats.get("completion_tokens", 0)
    ws_data["total_tokens"] += run_stats.get("total_tokens", 0)
    ws_data["total_cost_usd"] += run_stats.get("cost_usd", 0.0)
    ws_data["runs"].append(run_entry)
    _save_json_file(ws_file, ws_data)

    # Update global token file
    global_data = _load_json_file(_GLOBAL_TOKEN_FILE)
    if "total_calls" not in global_data:
        global_data = {
            "total_calls": 0, "total_prompt_tokens": 0,
            "total_completion_tokens": 0, "total_tokens": 0,
            "total_cost_usd": 0.0, "last_reset": None,
        }
    global_data["total_calls"] += run_stats.get("calls", 0)
    global_data["total_prompt_tokens"] += run_stats.get("prompt_tokens", 0)
    global_data["total_completion_tokens"] += run_stats.get("completion_tokens", 0)
    global_data["total_tokens"] += run_stats.get("total_tokens", 0)
    global_data["total_cost_usd"] += run_stats.get("cost_usd", 0.0)
    _save_json_file(_GLOBAL_TOKEN_FILE, global_data)


@app.get("/api/token-stats")
async def get_token_stats(workspace: str = Query("SIGIR25")):
    """Return token stats: current tracker, workspace, and global."""
    tracker = get_global_tracker()
    current_run = tracker.to_dict()

    ws_data = _load_json_file(_get_workspace_token_file(workspace))
    global_data = _load_json_file(_GLOBAL_TOKEN_FILE)

    return {
        "result": {
            "currentRun": current_run,
            "workspace": ws_data if ws_data else None,
            "global": global_data if global_data else None,
        }
    }


@app.get("/api/token-stats/workspace/{name}")
async def get_workspace_token_stats(name: str):
    """Return detailed token stats for a specific workspace."""
    ws_data = _load_json_file(_get_workspace_token_file(name))
    return {"result": ws_data if ws_data else {"workspace": name, "runs": []}}


@app.post("/api/token-stats/reset/workspace/{name}")
async def reset_workspace_token_stats(name: str):
    """Reset token stats for a workspace."""
    ws_file = _get_workspace_token_file(name)
    if ws_file.exists():
        ws_file.unlink()
    return {"success": True, "message": f"Token stats reset for workspace '{name}'."}


@app.post("/api/token-stats/reset/global")
async def reset_global_token_stats():
    """Reset global token stats."""
    import datetime
    _save_json_file(_GLOBAL_TOKEN_FILE, {
        "total_calls": 0, "total_prompt_tokens": 0,
        "total_completion_tokens": 0, "total_tokens": 0,
        "total_cost_usd": 0.0,
        "last_reset": datetime.datetime.now(datetime.timezone.utc).isoformat(),
    })
    return {"success": True, "message": "Global token stats reset."}


# ════════════════════════════════════════════════════════════════════
# API: Settings
# ════════════════════════════════════════════════════════════════════

# Track manually-entered API keys (session-only, never persisted to disk)
_manual_api_keys: dict[str, str] = {}

# Map provider -> environment variable name
_PROVIDER_ENV_KEYS = {
    "openrouter": "OPENROUTER_API_KEY",
}

@app.get("/api/models")
async def list_models(provider: str = None):
    """Fetch available models from OpenRouter API with pricing.

    Returns a list of model IDs grouped by provider, with per-model pricing.
    Requires OPENROUTER_API_KEY.
    """
    api_key = os.environ.get("OPENROUTER_API_KEY") or _manual_api_keys.get("openrouter")
    if not api_key:
        return {"success": False, "error": "No OpenRouter API key configured", "models": []}

    try:
        import httpx
        headers = {"Authorization": f"Bearer {api_key}"}
        resp = httpx.get("https://openrouter.ai/api/v1/models", headers=headers, timeout=15)
        data = resp.json().get("data", [])

        models_with_pricing = []
        for m in data:
            mid = m.get("id", "")
            # Filter to text-capable models
            arch = m.get("architecture", {})
            output_mods = arch.get("output_modalities", []) if arch else []
            if output_mods and "text" not in output_mods:
                continue
            pricing = m.get("pricing", {})
            prompt_price = float(pricing.get("prompt", "0") or "0")
            completion_price = float(pricing.get("completion", "0") or "0")
            models_with_pricing.append({
                "id": mid,
                "name": m.get("name", mid),
                "context_length": m.get("context_length", 0),
                "prompt_price_per_1m": round(prompt_price * 1_000_000, 4),
                "completion_price_per_1m": round(completion_price * 1_000_000, 4),
            })
        models_with_pricing.sort(key=lambda x: x["id"])

        # Populate live pricing cache for accurate cost estimation
        from token_tracker import set_live_pricing
        set_live_pricing(models_with_pricing)

        return {"success": True,
                "models": [m["id"] for m in models_with_pricing],
                "models_with_pricing": models_with_pricing}

        return {"success": True, "models": models}

    except Exception as e:
        logger.warning(f"Failed to list models for {prov}: {e}")
        return {"success": False, "error": str(e), "models": []}


@app.get("/api/settings")
async def get_settings():
    """Return current settings from the config module."""
    provider = getattr(config, "LLM_PROVIDER", "openai")
    env_key = _PROVIDER_ENV_KEYS.get(provider, "")
    has_key = bool(os.environ.get(env_key) or _manual_api_keys.get(provider))

    # Key status for all providers so UI can show status on provider switch
    api_keys_status = {
        p: bool(os.environ.get(ek) or _manual_api_keys.get(p))
        for p, ek in _PROVIDER_ENV_KEYS.items()
    }

    return {
        "result": {
            "llm": {
                "provider": provider,
                "model": getattr(config, "LLM_MODEL", "gpt-4o"),
                "temperature": getattr(config, "LLM_TEMPERATURE", 0.3),
                "api_key_source": "manual" if _manual_api_keys.get(provider) else "environment",
                "api_key_set": has_key,
                "api_keys_status": api_keys_status,
            },
            "oral": {
                "method": getattr(config, "ORAL_METHOD", "greedy"),
                "solver": getattr(config, "ORAL_SOLVER", "heuristic"),
                "alpha": getattr(config, "ORAL_ALPHA", 1.0),
                "enable_conflict_avoidance": getattr(config, "ENABLE_CONFLICT_AVOIDANCE", True),
            },
            "poster": {
                "method": getattr(config, "POSTER_METHOD", "greedy"),
                "solver": getattr(config, "POSTER_SOLVER", "heuristic"),
                "alpha": getattr(config, "POSTER_ALPHA", 1.0),
                "enable_conflict_avoidance": getattr(config, "POSTER_ENABLE_CONFLICT_AVOIDANCE", True),
                "proximity": getattr(config, "POSTER_PROXIMITY", True),
            },
            "similarity": {
                "method": getattr(config, "SIMILARITY_METHOD", "tfidf"),
                "embedding_model": getattr(config, "EMBEDDING_MODEL", "all-MiniLM-L6-v2"),
                "cache_enabled": getattr(config, "EMBEDDING_CACHE_ENABLED", True),
            },
        }
    }


@app.put("/api/settings")
async def update_settings(request: Request):
    """Update settings in-memory. Optionally accepts a manual API key (session-only)."""
    body = await request.json()

    llm = body.get("llm", {})
    if llm.get("provider"):
        config.LLM_PROVIDER = llm["provider"]
    if llm.get("model"):
        config.LLM_MODEL = llm["model"]
    if llm.get("temperature") is not None:
        config.LLM_TEMPERATURE = float(llm["temperature"])
    if llm.get("api_key"):
        # Store manual key in session memory AND set as env var so LLMClient picks it up
        provider = llm.get("provider", config.LLM_PROVIDER)
        env_key = _PROVIDER_ENV_KEYS.get(provider, "")
        if env_key:
            _manual_api_keys[provider] = llm["api_key"]
            os.environ[env_key] = llm["api_key"]
            logger.info(f"Manual API key set for provider '{provider}' (session-only)")

    oral = body.get("oral", {})
    if oral.get("method"):
        config.ORAL_METHOD = oral["method"]
    if oral.get("solver"):
        config.ORAL_SOLVER = oral["solver"]
    if oral.get("alpha") is not None:
        config.ORAL_ALPHA = float(oral["alpha"])
    if "enable_conflict_avoidance" in oral:
        config.ENABLE_CONFLICT_AVOIDANCE = bool(oral["enable_conflict_avoidance"])

    poster = body.get("poster", {})
    if poster.get("method"):
        config.POSTER_METHOD = poster["method"]
    if poster.get("solver"):
        config.POSTER_SOLVER = poster["solver"]
    if poster.get("alpha") is not None:
        config.POSTER_ALPHA = float(poster["alpha"])
    if "enable_conflict_avoidance" in poster:
        config.POSTER_ENABLE_CONFLICT_AVOIDANCE = bool(poster["enable_conflict_avoidance"])
    if "proximity" in poster:
        config.POSTER_PROXIMITY = bool(poster["proximity"])

    sim = body.get("similarity", {})
    if sim.get("method"):
        config.SIMILARITY_METHOD = sim["method"]
    if sim.get("embedding_model"):
        config.EMBEDDING_MODEL = sim["embedding_model"]
    if "cache_enabled" in sim:
        config.EMBEDDING_CACHE_ENABLED = bool(sim["cache_enabled"])

    # Clear caches so next run uses new settings
    _taxonomy_cache.clear()
    _similarity_cache.clear()

    logger.info(f"Settings updated: provider={config.LLM_PROVIDER}, model={config.LLM_MODEL}")
    return {"success": True, "message": "Settings updated."}


@app.post("/api/settings/test-llm")
async def test_llm_connection(request: Request):
    """Test LLM connectivity with current or provided settings."""
    body = await request.json()
    provider = body.get("provider", config.LLM_PROVIDER)
    model = body.get("model", config.LLM_MODEL)

    # If a manual key is provided, temporarily set it
    temp_key = body.get("api_key")
    env_key = _PROVIDER_ENV_KEYS.get(provider, "")
    old_env = None
    if temp_key and env_key:
        old_env = os.environ.get(env_key)
        os.environ[env_key] = temp_key

    # Temporarily override config for this test
    old_provider = config.LLM_PROVIDER
    old_model = config.LLM_MODEL
    config.LLM_PROVIDER = provider
    config.LLM_MODEL = model

    try:
        llm = LLMClient(json_mode=False)
        response = llm.chat("You are a helpful assistant.", "Say 'hello' in one word.", call_label="test-connection")
        return {"success": True, "message": f"Model responded: {response[:80]}"}
    except Exception as e:
        return {"success": False, "error": str(e)}
    finally:
        config.LLM_PROVIDER = old_provider
        config.LLM_MODEL = old_model
        if temp_key and env_key:
            if old_env is not None:
                os.environ[env_key] = old_env
            else:
                os.environ.pop(env_key, None)


# ════════════════════════════════════════════════════════════════════
# Catch-all static file route (MUST be after all /api/ routes)
# ════════════════════════════════════════════════════════════════════

@app.get("/{filepath:path}")
async def serve_static(filepath: str):
    """Serve static assets (CSS, JS, images). Registered last so API routes win."""
    # Strip query strings (e.g. ?v=2 cache busting)
    clean_path = filepath.split("?")[0] if "?" in filepath else filepath
    full_path = PROJECT_ROOT / clean_path
    if full_path.is_file() and full_path.suffix in STATIC_EXTENSIONS:
        resp = FileResponse(str(full_path))
        # Prevent aggressive caching of JS/CSS during development
        if full_path.suffix in (".js", ".css"):
            resp.headers["Cache-Control"] = "no-cache, must-revalidate"
        return resp
    return JSONResponse({"error": "Not found"}, status_code=404)


# ════════════════════════════════════════════════════════════════════
# Entrypoint
# ════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="TaxoConf API Server")
    parser.add_argument("--host", default="127.0.0.1", help="Bind host")
    parser.add_argument("--port", type=int, default=8000, help="Port")
    parser.add_argument("--reload", action="store_true", help="Auto-reload on code changes")
    parser.add_argument("--config", type=str, default=None, help="YAML config file")
    args = parser.parse_args()

    # Load YAML config if provided
    if args.config:
        config.load_from_yaml(args.config)

    conferences = discover_conferences()
    logger.info(f"Available conferences: {conferences}")
    logger.info(f"Starting TaxoConf server on http://{args.host}:{args.port}")

    uvicorn.run(
        "server:app",
        host=args.host,
        port=args.port,
        reload=args.reload,
        log_level="info",
    )


if __name__ == "__main__":
    main()
